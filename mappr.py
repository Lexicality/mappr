from __future__ import generator_stop

import typing  # noqa: F401
from typing import Dict, Optional, Set, Iterable, Iterator

from openpyxl import load_workbook
from openpyxl.cell import Cell as xlCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import Worksheet as xlSheet
# from openpyxl.styles import Color, colors

cellcache: Dict[str, 'Cell'] = {}


def _valid_coord(sheet: xlSheet, x: int, y: int) -> bool:
    # Note that column 1 is the data column
    return (2 <= x and x <= sheet.max_column) and (1 <= y and y <= sheet.max_row)


class Cell:
    class Coords(Iterable):
        def __init__(self, parent: 'Cell') -> None:
            self.cell = parent

        @property
        def N(self):
            return self.cell._get_neighbour(0, -1)

        @property
        def E(self):
            return self.cell._get_neighbour(1, 0)

        @property
        def S(self):
            return self.cell._get_neighbour(0, 1)

        @property
        def W(self):
            return self.cell._get_neighbour(-1, 0)

        @property
        def NE(self):
            return self.cell._get_neighbour(1, -1)

        @property
        def SE(self):
            return self.cell._get_neighbour(1, 1)

        @property
        def SW(self):
            return self.cell._get_neighbour(-1, 1)

        @property
        def NW(self):
            return self.cell._get_neighbour(-1, -1)

        def __iter__(self) -> Iterator['Cell']:
            return iter([self.N, self.NE, self.E, self.SE, self.S, self.SW, self.W, self.NW])

    def __init__(self, cell: xlCell) -> None:
        self.cell = cell
        self.neighbours = Cell.Coords(self)

    def is_void(self) -> bool:
        return (
            not self.cell.has_style or
            self.cell.fill.bgColor == 'FFFFFFFF' or
            self.cell.fill.bgColor == '00FFFFFF'
        )

    def is_me(self, other: 'Cell') -> bool:
        return self.cell.fill.bgColor == other.cell.fill.bgColor

    def get_all_me(self) -> Set['Cell']:
        to_check: Set[Cell] = set([self])
        found: Set[Cell] = set([self])
        has_checked: Set[Cell] = set()
        while to_check:
            cell = to_check.pop()
            has_checked.add(cell)
            for n in cell.neighbours:
                if n and n not in has_checked and self.is_me(n):
                    found.add(n)
                    to_check.add(n)

        return found

    def _get_neighbour(self, x: int, y: int) -> Optional['Cell']:
        col = self.cell.col_idx + x
        row = self.cell.row + y
        if not _valid_coord(self.cell.parent, row, col):
            return None
        coords = "{col}{row}".format(col=get_column_letter(col), row=row)
        c = cellcache.get(coords)
        if not c:
            c = Cell(self.cell.parent[coords])
            cellcache[coords] = c
        return c


wb = load_workbook('test.xlsx')

sheet = wb.active
cell = sheet['H4']
mycell = Cell(cell)
